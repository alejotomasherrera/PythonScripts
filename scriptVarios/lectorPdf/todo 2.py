from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import unicodedata
from PyPDF2 import PdfReader
import pandas as pd
import re

def extraer_texto(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PdfReader(file)
        text = ""
        for page in reader.pages:
            text += page.extract_text()
        return text


def guardar_en_excel(datos, nombre_archivo):
    # Si 'datos' es un diccionario de valores escalares, conviértelo en un diccionario con una lista
    for key in datos:
        datos[key] = [datos[key]]

    df = pd.DataFrame(datos)
    df.to_excel(nombre_archivo, index=False)

# Suponiendo que 'detalles_operativos' es una lista de diccionarios con la información de cada detalle operativo
def guardar_detalle_en_excel(detalles_operativos, nombre_archivo, id_parte_diario):
    # Crear un DataFrame con los detalles operativos
    df_detalles = pd.DataFrame(detalles_operativos)

    # Agregar la columna idParteDiarioPerforacion
    df_detalles['idParteDiarioPerforacion'] = id_parte_diario

    # Crear un writer de pandas y guardar la información en una nueva hoja
    with pd.ExcelWriter(nombre_archivo, mode='a', engine='openpyxl') as writer:
        df_detalles.to_excel(writer, sheet_name='DetalleParteDiario', index=False)

def limpiar_yacimiento(yacimiento):
    palabra_no_deseada = 'Supervisor'
    if palabra_no_deseada in yacimiento:
        # Encuentra la posición donde comienza la palabra no deseada y corta el string hasta ese punto.
        yacimiento = yacimiento[:yacimiento.index(palabra_no_deseada)].strip()
    return yacimiento

def procesar_texto(texto):
    # Ajusta estas expresiones regulares según sea necesario
    fecha_reporte = re.search(r'Fecha\s*Reporte: \s*(\d{2}/\d{2}/\d{4})', texto)
    fecha_reporte = fecha_reporte.group(1) if fecha_reporte else 'No encontrada'

    yacimiento = re.search(r'Yacimiento:\s*([\w\s]+)', texto)
    yacimiento = limpiar_yacimiento(yacimiento.group(1)) if yacimiento else 'No encontrado'

    # Ajusta esto si hay espacios u otros caracteres entre 'AFE' y 'No'
    afe_no = re.search(r'AFE\s*No.?:\s*(\S+)', texto)
    afe_no = afe_no.group(1) if afe_no else 'No encontrado'
    afe_no = afe_no.replace('Fecha', '')  # Elimina la palabra 'Fecha' si está presente

    return {
        'Fecha Reporte': fecha_reporte,
        'Yacimiento': yacimiento,
        'AFE No': afe_no.strip()  # Elimina espacios en blanco adicionales
    }

# Asegúrate de cambiar "petroleo1.pdf" al nombre de tu archivo PDF NOMBRE DEL PDF
#--------------------------------------------------------------------------------
texto_pdf = extraer_texto("Partes Diarios/5.pdf")
datos = procesar_texto(texto_pdf)
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
xlsx_path = 'partediarioprocesado.xlsx'
guardar_en_excel(datos, xlsx_path)
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# ESTE ES EL TEXTO A PEGAR DEL DETALLE -----------------------------------------------------------------------------------------------------------------------------
# Texto crudo proporcionado por el usuario.
raw_text = """
Desde Hasta Duración Fase Código Resúmen Operativo
06:00 08:15 2,25 P-ENTUBMENTUB Entuba pozo con Csg de 5" (21.4 Lb/ft, P-110, Wedge 625), de 6250 m a 6504.05 m con rotación 30 rpm, 15/16klbs.ft Controla desplazamiento en TT alineado por MPD. 08:15 08:30 0,25 P-ENTUBBUILDUP BU. Pozo sin presion.
08:30 09:00 0,50 P-ENTUBFCHECK Flow check, pozo estático.
09:00 09:30 0,50 P-ENTUBDBEARING Desmonta BA con ultimo casing.
09:30 10:00 0,50 P-ENTUBCUELGA Personal de Cia MMA levanta cjto de colgador mandrel de 5", bajante vinculado y niple de maniobra 5". Conecta y torquea colgador mandrel con Csg.Profundiza, con rotación, y cuelga Csg.*Profundidad de zapato: 6502.55 m.*Profundidad de collar: 6490.16 m.
10:00 10:15 0,25 P-CEMENCEMENT Llena directa y corrobora circulación. Desmonta casing de maniobra.Fin de evento.
10:15 12:30 2,25 DTM-PRE CABEZA* Lava alojamiento con gasoil. Empaqueta pozo, realiza PH y prueba de tensión. OK.
12:30 13:45 1,25 DTM-PRE DCRT* Desmonta CRT + servicio de entubación.
13:45 15:00 1,25 DTM-PRE DMPD* Reunión seg y operatiba con cia WTFD. Monta tapon en RCD. Purga lineas de MPD.
15:00 17:00 2,00 DTM-PRE SK/DARMA Equipo purga y desmonta lineas de alta, ignifugo, manifold, stand pipe y bbas. Retira bandeja ecologica.
17:00 18:30 1,50 DTM-PRE SK/DARMACia WTFD MPD desarma linea primaria, secundaria. Saca lonas de RCD,
18:30 20:30 2,00 DTM-PRE DMPD* Lava BOP parcialmente. Posiciona y acondiciona en boomer con amelas y vincula a BOP sobre anular. Retira salida lateral.
20:30 23:15 2,75 DTM-PRE SK/DARMAAcondiciona para realizar Skid, coloca vigas, retira parrillas ante pozos. acondiciona manguerotes, retira escalera.barandas de flow line. Retiran mantas de bodega. Instala mangueras en carretel de cable perf.
23:15 23:30 0,25 DTM-PRE SK/DARMARealiza charla previa a realizar Skidding.
23:30 02:00 2,50 DTM-PRE SK/DARMARealiza Skid de FP-1468(h) a FP-1465(h), sobre 30 m, para desarmar sondeo a playa.
02:00 02:30 0,50 DTM-PRE SK/DARMAAsienta stack BOP, cia MMA coloca prisioneros CR-200. Acondiciona post skid (escaleras y mangueras de BOP).
02:30 02:45 0,25 DTM-PRE SK/DARMACia WTFD MPD realiza charla seg y operativa para desmontar equipos.
02:45 06:00 3,25 DTM-PREDMPD* MPD desmonta equipos, líneas y RCD, en progreso.En simultaneo, limpia bodega en FP-1468(h) y cia MMA coloca brida combinadora 13.5/8" x 5.1/8"-15k + válvmaestra 5.1/8" + brida 5.1/8" x 5 TXP para cabeza cementación- Realiza pruebas, en progreso.
"""
# ESTE ES EL TEXTO A PEGAR DEL DETALLE -----------------------------------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Patrones para las columnas
patron_fila = re.compile(r'(\d{2}:\d{2})\s(\d{2}:\d{2})\s([\d,]+)\s(\S+)\s(\S+)\s(.+?)(?=\s\d{2}:\d{2}|\Z)', re.DOTALL)

# Encontramos todas las coincidencias del patrón en el texto
matches = patron_fila.findall(raw_text)

# Creamos un DataFrame con las coincidencias
df = pd.DataFrame(matches, columns=['Desde', 'Hasta', 'Duración', 'Fase', 'Código', 'Resumen Operativo'])

# Convertimos las comas en puntos en la columna 'Duración'
df['Duración'] = df['Duración'].str.replace(',', '.')

# Normalizamos los caracteres
df['Resumen Operativo'] = df['Resumen Operativo'].apply(
    lambda x: unicodedata.normalize('NFKD', x).encode('ascii', 'ignore').decode('ascii')
)
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Exportamos el DataFrame a una hoja específica en un archivo Excel
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Intentamos cargar el libro de trabajo, si falla, creamos uno nuevo
try:
    book = load_workbook(xlsx_path)
    # Si la hoja existe, la eliminamos
    if 'detalleresumen' in book.sheetnames:
        book.remove(book['detalleresumen'])
except Exception as e:
    print(f"Error al cargar el archivo: {e}. Creando un nuevo libro de trabajo.")
    book = Workbook()

# Añadimos una nueva hoja
sheet = book.create_sheet('detalleresumen')

# Convertimos el DataFrame en filas para la hoja de Excel
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        sheet.cell(row=r_idx, column=c_idx, value=value)

# Guardamos el libro de trabajo
book.save(xlsx_path)

print(f"Los datos han sido añadidos a la hoja 'detalleresumen' en el archivo Excel: {xlsx_path}")
